# -*- coding: utf-8 -*-
import os, sys, sqlite3, json, re
from pathlib import Path
try:
    import openpyxl
except Exception:
    print("[skip] openpyxl fehlt – XLSX wird übersprungen")
    sys.exit(0)

DB = Path("data/manifest.db")
ASSETS = "assets"; FACTS = "facts"

def upsert_asset(cur, path, customer="(unknown)", section="kpis", tags="[]", atype="xlsx"):
    cur.execute(f"SELECT id FROM {ASSETS} WHERE path=?",(path,))
    row = cur.fetchone()
    if row: return row[0]
    cur.execute(f"""INSERT INTO {ASSETS}
        (customer,path,type,size,hash,tags,section_hint,topic_hint,source,created_at)
        VALUES(?,?,?,?,?,?,?,?,?,strftime('%s','now'))""",
        (customer, path, atype, os.path.getsize(path) if os.path.exists(path) else 0,
         None, tags, section, None, "xlsx_to_facts"))
    return cur.lastrowid

def insert_fact(cur, asset_id, metric, value, unit=None, period=None, cite=None):
    cur.execute(f"""INSERT INTO {FACTS}
        (asset_id,metric,value,unit,period_start,period_end,citation)
        VALUES(?,?,?,?,?,?,?)""",
        (asset_id, metric, value, unit, period, None, cite))

def normalize_header(v:str)->str:
    v=(v or "").strip().lower()
    v=v.replace("€"," eur ").replace("$"," usd ").replace("%"," pct ")
    v=re.sub(r"\s+"," ",v)
    return v

def metric_from_header(h:str):
    """
    Wandelt Header in (metric, unit_hint) um.
    z.B. "ctr pct" -> ("ctr","%"), "cpc eur" -> ("cpc","eur")
    """
    h=normalize_header(h)
    if h.startswith("ctr"):   return ("ctr","%")
    if h.startswith("cpc"):   return ("cpc","eur" if "eur" in h else None)
    if h.startswith("cac"):   return ("cac","eur" if "eur" in h else None)
    if h.startswith("roas"):  return ("roas",None)
    return (h.split(" ")[0], None)

def main(paths):
    if not DB.exists():
        print("[ERR] manifest.db fehlt – bitte zuerst Wave 1 laufen lassen", file=sys.stderr); sys.exit(1)
    con=sqlite3.connect(DB); cur=con.cursor()
    for p in paths:
        p=Path(p)
        if not p.exists():
            print(f"[skip] nicht gefunden: {p}")
            continue
        wb=openpyxl.load_workbook(p, data_only=True)
        ws=wb.active

        # Header suchen (erste Zeile mit Textzellen)
        header=[]
        for r in ws.iter_rows(min_row=1, max_row=10):
            vals=[(c.value if c.value is not None else "") for c in r]
            if any(str(v).strip() for v in vals):
                header=[str(v) if v is not None else "" for v in vals]
                break
        if not header:
            print(f"[skip] keine Header gefunden in {p.name}")
            continue

        hix={i:normalize_header(h) for i,h in enumerate(header)}
        asset_id = upsert_asset(cur, str(p), section="kpis", atype="xlsx")
        rows=0

        # Modus A: long-form
        wants = {v:i for i,v in hix.items()}
        if "metric" in wants and "value" in wants:
            per_key = ("period" if "period" in wants else
                       ("quarter" if "quarter" in wants else
                        ("month" if "month" in wants else None)))
            unit_key = "unit" if "unit" in wants else None
            for r in ws.iter_rows(min_row=2, values_only=True):
                m_raw = str(r[wants["metric"]]).strip().lower() if r[wants["metric"]] is not None else ""
                if not m_raw: continue
                metric, unit_hint = metric_from_header(m_raw)
                val=r[wants["value"]]
                try: value=float(str(val).replace(",",".")); 
                except: continue
                period=(str(r[wants[per_key]]).strip() if per_key else None)
                unit=(str(r[wants[unit_key]]).strip() if unit_key else unit_hint)
                insert_fact(cur, asset_id, metric, value, unit, period, cite=str(p.name))
                rows+=1
        else:
            # Modus B: wide-form – erste Spalte = period
            per_idx = None
            for i,h in hix.items():
                if h.startswith(("period","quarter","date")): 
                    per_idx=i; break
            if per_idx is None: per_idx=0
            for r in ws.iter_rows(min_row=2, values_only=True):
                period = r[per_idx]
                period = str(period).strip() if period is not None else None
                for i,h in hix.items():
                    if i==per_idx or not h: continue
                    try:
                        cell=r[i]
                        if cell is None: continue
                        v=float(str(cell).replace(",",".")); 
                        metric, unit_hint = metric_from_header(header[i])
                        unit = unit_hint
                        insert_fact(cur, asset_id, metric, v, unit, period, cite=str(p.name))
                        rows+=1
                    except:
                        pass
        con.commit()
        print(f"[ok] {p.name}: {rows} facts geschrieben")
    con.close()

if __name__=="__main__":
    main(sys.argv[1:] or [])
